#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_data.py - Génère le Portail Tech PWA mono-fichier
Lit un classeur RPM.xlsm → produit index.html + met à jour sw.js

Usage :
    python export_data.py "chemin/RPM.xlsm" [dossier_sortie]

Placeholders remplacés dans template.html :
    %%DATA_ETALONS%%     JSON étalons
    %%DATA_CATEGORIES%%  JSON catégories
    %%DATA_MATERIELS%%   JSON matériels
    %%DATA_CONSTATS%%    JSON fiches constats (scan forms/)
    %%DATE_ETALONS%%     Date MAJ étalons
    %%DATE_MATERIELS%%   Date MAJ matériels
    %%DATE_CONSTATS%%    Date MAJ constats
    %%NB_ETALONS%%       Nombre étalons
    %%NB_MATERIELS%%     Nombre matériels
    %%NB_CONSTATS%%      Nombre fiches constats
    %%ETIQ_IMG_C%%       Image base64 étiquette conforme
    %%ETIQ_IMG_NC%%      Image base64 étiquette non conforme

Placeholders remplacés dans sw.js :
    %%SW_VERSION%%       Version auto-incrémentée
    %%SW_FILES%%         Liste des forms/*.html à cacher
"""

import sys, os, json, datetime, re, hashlib
from pathlib import Path
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("ERREUR : pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════
CATEGORIES_DEFAUT = {
    "Source": ["Demi-vie (jours)", "Isotope", "Date référence", "Activité ref (Bq)", "Type EDF"],
    "Irradiateur": ["Distance 1", "Distance 2"],
    "Débitmètre": [],
    "Banc test MIP-ECM": [],
    "Oscilloscope": [],
    "Multimètre": ["Marque/Modèle", "N° série"],
}

# Fichier de suivi des dates de MAJ par section
DATES_FILE = ".portail_dates.json"


# ══════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════
def fmt_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d %H:%M") if val.hour or val.minute else val.strftime("%Y-%m-%d")
    return str(val).strip()


def fmt_val(val):
    if val is None:
        return ""
    return str(val).strip()


def norm_site(s):
    if not s:
        return ""
    return s.strip().title()


def statut_valid(ds):
    if not ds:
        return "inconnu"
    try:
        d = datetime.datetime.strptime(ds[:10], "%Y-%m-%d").date()
    except Exception:
        return "inconnu"
    if d.year >= 2999:
        return "permanent"
    delta = (d - datetime.date.today()).days
    if delta < 0:
        return "expire"
    elif delta <= 90:
        return "bientot"
    return "valide"


def data_hash(data):
    """Hash JSON pour détecter les changements."""
    return hashlib.md5(json.dumps(data, sort_keys=True, ensure_ascii=False).encode()).hexdigest()[:12]


def load_dates(dossier):
    """Charge les dates de MAJ précédentes."""
    p = dossier / DATES_FILE
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_dates(dossier, dates):
    """Sauvegarde les dates de MAJ."""
    p = dossier / DATES_FILE
    p.write_text(json.dumps(dates, ensure_ascii=False, indent=2), encoding="utf-8")


# ══════════════════════════════════════════════════════════
# EXPORT ÉTALONS
# ══════════════════════════════════════════════════════════
def export_etalons(wb):
    if "Etalons" not in wb.sheetnames:
        return []
    ws = wb["Etalons"]
    out = []
    for row in ws.iter_rows(min_row=2, max_col=26, values_only=True):
        if not row[0]:
            continue
        vs = fmt_date(row[2])
        cat = fmt_val(row[1])
        cat_champs = CATEGORIES_DEFAUT.get(cat, [])
        champs = []
        for i in range(15):
            v = fmt_val(row[5 + i])
            if v:
                l = cat_champs[i] if i < len(cat_champs) and cat_champs[i] else f"Champ {i+1}"
                champs.append({"l": l, "v": v})
        src = {}
        if cat == "Source":
            try:
                src["tv"] = float(fmt_val(row[5]).replace(",", ".").replace(" ", ""))
            except Exception:
                pass
            try:
                src["a0"] = float(fmt_val(row[8]).replace(",", ".").replace(" ", ""))
            except Exception:
                pass
            dr = row[7]
            if isinstance(dr, datetime.datetime):
                src["dr"] = dr.strftime("%d/%m/%Y")
            elif dr:
                src["dr"] = fmt_val(dr)
        out.append({
            "id": fmt_val(row[0]), "cat": cat, "val": vs, "sv": statut_valid(vs),
            "site": norm_site(fmt_val(row[3])), "loc": fmt_val(row[4]),
            "ch": champs, "src": src,
            "dc": fmt_date(row[20]), "dm": fmt_date(row[21]),
            "uc": fmt_val(row[22]), "um": fmt_val(row[23]),
        })
    return out


# ══════════════════════════════════════════════════════════
# EXPORT MATÉRIELS
# ══════════════════════════════════════════════════════════
def export_materiels(wb):
    if "Materiels" not in wb.sheetnames:
        return []
    ws = wb["Materiels"]
    out = []
    for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        if not row[0]:
            continue
        vs = fmt_date(row[5])
        out.append({
            "id": fmt_val(row[0]), "des": fmt_val(row[1]),
            "ns": fmt_val(row[2]), "nc": fmt_val(row[3]),
            "dvp": fmt_date(row[4]), "val": vs, "sv": statut_valid(vs),
            "cnpe": norm_site(fmt_val(row[6])), "loc": fmt_val(row[7]),
            "ctr": fmt_val(row[8]), "obs": fmt_val(row[9])
        })
    return out


# ══════════════════════════════════════════════════════════
# SCAN FICHES CONSTATS (forms/*.html)
# ══════════════════════════════════════════════════════════
def scanner_forms(dossier_sortie):
    forms_dir = dossier_sortie / "forms"
    fiches = []
    if not forms_dir.exists():
        forms_dir.mkdir(exist_ok=True)
        return fiches
    for f in sorted(forms_dir.glob("*.html")):
        meta = {}
        try:
            contenu = f.read_text(encoding="utf-8", errors="ignore")[:4000]
            for b in re.finditer(
                r'<meta\s+name=["\']fiche-(\w+)["\']\s+content=["\']([^"\']*)["\']',
                contenu, re.IGNORECASE
            ):
                meta[b.group(1)] = b.group(2)
        except Exception:
            pass
        nom = f.stem.replace("_", " ").replace("-", " ").title()
        fiches.append({
            "f": f.name,
            "url": "forms/" + f.name,
            "t": meta.get("titre", nom),
            "d": meta.get("description", "Formulaire de vérification"),
            "c": meta.get("categorie", "Général"),
            "i": meta.get("icon", "📋"),
        })
    return fiches


# ══════════════════════════════════════════════════════════
# IMAGES ÉTIQUETTES BASE64
# ══════════════════════════════════════════════════════════
def load_etiq_images(dossier_script):
    """Charge les images base64 depuis les fichiers JPEG du dossier."""
    images = {"C": "", "NC": ""}
    for name, key in [("etiq_conforme.jpg", "C"), ("etiq_non_conforme.jpg", "NC")]:
        p = dossier_script / name
        if p.exists():
            import base64
            b64 = base64.b64encode(p.read_bytes()).decode()
            images[key] = f"data:image/jpeg;base64,{b64}"
    return images


# ══════════════════════════════════════════════════════════
# MISE À JOUR SW.JS
# ══════════════════════════════════════════════════════════
def update_sw(dossier_sortie, fiches, version):
    """Met à jour sw.js avec la liste des fiches et la version."""
    sw_template = Path(__file__).parent / "sw.js"
    if not sw_template.exists():
        print("  ⚠ sw.js template introuvable")
        return

    sw_content = sw_template.read_text(encoding="utf-8")

    # Construire la liste des forms à cacher
    forms_lines = []
    for f in fiches:
        forms_lines.append(f"  './forms/{f['f']}',")
    sw_files = "\n".join(forms_lines) if forms_lines else "  // (aucune fiche)"

    sw_content = sw_content.replace("%%SW_VERSION%%", version)
    sw_content = sw_content.replace("%%SW_FILES%%", sw_files)

    sw_out = dossier_sortie / "sw.js"
    sw_out.write_text(sw_content, encoding="utf-8")
    print(f"  ✓ sw.js mis à jour (v{version}, {len(fiches)} fiche(s) en cache)")


# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════
def main():
    if len(sys.argv) < 2:
        print("Usage : python export_data.py <fichier_rpm.xlsm> [dossier_sortie]")
        sys.exit(1)

    chemin_rpm = Path(sys.argv[1])
    dossier_sortie = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).parent
    dossier_script = Path(__file__).parent

    if not chemin_rpm.exists():
        print(f"ERREUR : Fichier introuvable : {chemin_rpm}")
        sys.exit(1)
    dossier_sortie.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*55}")
    print(f"  ☢  Export Portail Tech PWA")
    print(f"  Source : {chemin_rpm.name}")
    print(f"{'='*55}\n")

    # ── Charger template ──────────────────────────────────
    template_path = dossier_script / "template.html"
    if not template_path.exists():
        print("ERREUR : template.html introuvable !")
        sys.exit(1)
    html = template_path.read_text(encoding="utf-8")

    # ── Charger dates précédentes ─────────────────────────
    dates = load_dates(dossier_sortie)
    now = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── Lire le classeur RPM ──────────────────────────────
    wb = openpyxl.load_workbook(str(chemin_rpm), read_only=True, data_only=True)

    # ── Étalons ───────────────────────────────────────────
    print("  📦 Étalons...")
    etalons = export_etalons(wb)
    for c, n in Counter(e["cat"] for e in etalons).most_common():
        print(f"      • {c}: {n}")
    print(f"  ✓ {len(etalons)} étalon(s)")

    h_et = data_hash(etalons)
    if h_et != dates.get("hash_etalons"):
        dates["date_etalons"] = now
        dates["hash_etalons"] = h_et
        print(f"  → Données modifiées, MAJ: {now}")
    else:
        print(f"  → Inchangé depuis {dates.get('date_etalons', '?')}")

    # ── Matériels ─────────────────────────────────────────
    print("\n  📦 Matériels...")
    materiels = export_materiels(wb)
    for c, n in Counter(m["cnpe"] for m in materiels).most_common():
        print(f"      • {c or '(vide)'}: {n}")
    print(f"  ✓ {len(materiels)} matériel(s)")

    h_mat = data_hash(materiels)
    if h_mat != dates.get("hash_materiels"):
        dates["date_materiels"] = now
        dates["hash_materiels"] = h_mat
        print(f"  → Données modifiées, MAJ: {now}")
    else:
        print(f"  → Inchangé depuis {dates.get('date_materiels', '?')}")

    wb.close()

    # ── Constats (forms/) ─────────────────────────────────
    print("\n  📦 Constats (forms/)...")
    fiches = scanner_forms(dossier_sortie)
    if fiches:
        for f in fiches:
            print(f"      • {f['t']} ({f['f']})")
        print(f"  ✓ {len(fiches)} fiche(s)")
    else:
        print("  ⚠ Aucune fiche dans forms/")

    h_cst = data_hash(fiches)
    if h_cst != dates.get("hash_constats"):
        dates["date_constats"] = now
        dates["hash_constats"] = h_cst
        print(f"  → Fiches modifiées, MAJ: {now}")
    else:
        print(f"  → Inchangé depuis {dates.get('date_constats', '?')}")

    # ── Catégories par défaut ─────────────────────────────
    cats = [{"nom": k, "champs": v} for k, v in CATEGORIES_DEFAUT.items()]

    # ── Images étiquettes ─────────────────────────────────
    print("\n  🏷 Images étiquettes...")
    etiq_imgs = load_etiq_images(dossier_script)
    if etiq_imgs["C"]:
        print(f"      ✓ Conforme ({len(etiq_imgs['C'])//1024} Ko)")
    else:
        print("      ⚠ etiq_conforme.jpg introuvable (placeholder vide)")
    if etiq_imgs["NC"]:
        print(f"      ✓ Non conforme ({len(etiq_imgs['NC'])//1024} Ko)")
    else:
        print("      ⚠ etiq_non_conforme.jpg introuvable (placeholder vide)")

    # ══════════════════════════════════════════════════════
    # GÉNÉRATION index.html
    # ══════════════════════════════════════════════════════
    print("\n  🔧 Génération index.html...")
    html = html.replace('%%DATA_ETALONS%%', json.dumps(etalons, ensure_ascii=False, separators=(',', ':')))
    html = html.replace('%%DATA_MATERIELS%%', json.dumps(materiels, ensure_ascii=False, separators=(',', ':')))
    html = html.replace('%%DATA_CONSTATS%%', json.dumps(fiches, ensure_ascii=False, separators=(',', ':')))
    html = html.replace('%%DATA_CATEGORIES%%', json.dumps(cats, ensure_ascii=False, separators=(',', ':')))
    html = html.replace('%%DATE_ETALONS%%', dates.get("date_etalons", now))
    html = html.replace('%%DATE_MATERIELS%%', dates.get("date_materiels", now))
    html = html.replace('%%DATE_CONSTATS%%', dates.get("date_constats", now))
    html = html.replace('%%NB_ETALONS%%', str(len(etalons)))
    html = html.replace('%%NB_MATERIELS%%', str(len(materiels)))
    html = html.replace('%%NB_CONSTATS%%', str(len(fiches)))
    html = html.replace('%%ETIQ_IMG_C%%', etiq_imgs["C"])
    html = html.replace('%%ETIQ_IMG_NC%%', etiq_imgs["NC"])

    out_path = dossier_sortie / "index.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"  ✓ index.html ({out_path.stat().st_size / 1024:.0f} Ko)")

    # ══════════════════════════════════════════════════════
    # INJECTION SOURCES dans forms/*.html
    # ══════════════════════════════════════════════════════
    sources_json = json.dumps(
        [e for e in etalons if e["cat"] == "Source"],
        ensure_ascii=False, separators=(',', ':')
    )
    forms_dir = dossier_sortie / "forms"
    if forms_dir.exists():
        injected = 0
        for f in forms_dir.glob("*.html"):
            contenu = f.read_text(encoding="utf-8")
            if '%%DATA_SOURCES%%' in contenu:
                contenu = contenu.replace('%%DATA_SOURCES%%', sources_json)
                f.write_text(contenu, encoding="utf-8")
                injected += 1
                print(f"  ✓ Sources injectées dans {f.name}")
        if injected:
            nb_src = len([e for e in etalons if e["cat"] == "Source"])
            print(f"  → {nb_src} sources dans {injected} fiche(s)")

    # ══════════════════════════════════════════════════════
    # MISE À JOUR sw.js
    # ══════════════════════════════════════════════════════
    print("\n  🔧 Mise à jour sw.js...")
    # Version = hash court du contenu pour forcer la MàJ du cache
    sw_version = data_hash({"e": h_et, "m": h_mat, "c": h_cst, "t": now})
    update_sw(dossier_sortie, fiches, sw_version)

    # ══════════════════════════════════════════════════════
    # COPIE manifest.json
    # ══════════════════════════════════════════════════════
    manifest_src = dossier_script / "manifest.json"
    manifest_dst = dossier_sortie / "manifest.json"
    if manifest_src.exists() and str(manifest_src) != str(manifest_dst):
        manifest_dst.write_text(manifest_src.read_text(encoding="utf-8"), encoding="utf-8")
        print("  ✓ manifest.json copié")

    # ══════════════════════════════════════════════════════
    # SAUVEGARDE DATES
    # ══════════════════════════════════════════════════════
    save_dates(dossier_sortie, dates)

    # ── Résumé ────────────────────────────────────────────
    print(f"\n{'='*55}")
    print(f"  ✅ Export terminé !")
    print(f"     → {out_path}")
    print(f"     Étalons:   {len(etalons):>5}  (MAJ: {dates.get('date_etalons', '?')})")
    print(f"     Matériels: {len(materiels):>5}  (MAJ: {dates.get('date_materiels', '?')})")
    print(f"     Constats:  {len(fiches):>5}  (MAJ: {dates.get('date_constats', '?')})")
    print(f"     SW:         v{sw_version}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
